#!/usr/bin/env python3
"""
리포트 생성기
"""
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional
import psycopg2
from psycopg2.extras import RealDictCursor
import json
import os
from jinja2 import Environment, FileSystemLoader
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
import logging

logger = logging.getLogger(__name__)

# 한글 폰트 등록 (ReportLab)
try:
    pdfmetrics.registerFont(TTFont('NanumGothic', '/usr/share/fonts/truetype/nanum/NanumGothic.ttf'))
except:
    logger.warning("한글 폰트를 찾을 수 없습니다. 기본 폰트를 사용합니다.")


class ReportGenerator:
    """
    다양한 형식의 리포트를 생성하는 클래스
    """
    
    def __init__(self, db_config: Dict[str, Any]):
        self.db_config = db_config
        self.output_dir = "reports/output"
        os.makedirs(self.output_dir, exist_ok=True)
        
        # 스타일 설정
        self.styles = getSampleStyleSheet()
        self.styles.add(ParagraphStyle(
            name='KoreanStyle',
            fontName='NanumGothic',
            fontSize=10,
            leading=12
        ))
        
        # 차트 스타일 설정
        sns.set_style("whitegrid")
        plt.rcParams['font.family'] = 'NanumGothic'
        plt.rcParams['axes.unicode_minus'] = False
    
    def generate_sales_report(self, start_date: str, end_date: str, 
                            format: str = 'pdf', include_charts: bool = True) -> str:
        """매출 리포트 생성"""
        conn = psycopg2.connect(**self.db_config)
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # 1. 전체 매출 통계
        cursor.execute("""
            SELECT 
                COUNT(DISTINCT id) as total_orders,
                SUM(total_amount) as total_revenue,
                AVG(total_amount) as avg_order_value,
                COUNT(DISTINCT customer_id) as unique_customers
            FROM orders
            WHERE created_at BETWEEN %s AND %s
        """, (start_date, end_date))
        summary = cursor.fetchone()
        
        # 2. 일별 매출
        cursor.execute("""
            SELECT 
                DATE(created_at) as date,
                COUNT(*) as order_count,
                SUM(total_amount) as revenue
            FROM orders
            WHERE created_at BETWEEN %s AND %s
            GROUP BY DATE(created_at)
            ORDER BY date
        """, (start_date, end_date))
        daily_sales = cursor.fetchall()
        
        # 3. 카테고리별 매출
        cursor.execute("""
            SELECT 
                p.category,
                COUNT(DISTINCT o.id) as order_count,
                SUM(oi.quantity) as quantity_sold,
                SUM(oi.quantity * oi.price) as revenue
            FROM orders o
            JOIN order_items oi ON o.id = oi.order_id
            JOIN products p ON oi.product_id = p.id
            WHERE o.created_at BETWEEN %s AND %s
            GROUP BY p.category
            ORDER BY revenue DESC
        """, (start_date, end_date))
        category_sales = cursor.fetchall()
        
        # 4. 베스트셀러 상품
        cursor.execute("""
            SELECT 
                p.id,
                p.name,
                p.category,
                SUM(oi.quantity) as quantity_sold,
                SUM(oi.quantity * oi.price) as revenue
            FROM orders o
            JOIN order_items oi ON o.id = oi.order_id
            JOIN products p ON oi.product_id = p.id
            WHERE o.created_at BETWEEN %s AND %s
            GROUP BY p.id, p.name, p.category
            ORDER BY revenue DESC
            LIMIT 10
        """, (start_date, end_date))
        best_sellers = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        # 차트 생성
        charts = []
        if include_charts:
            charts = self._create_sales_charts(daily_sales, category_sales)
        
        # 리포트 생성
        report_data = {
            'title': '매출 분석 리포트',
            'period': f"{start_date} ~ {end_date}",
            'summary': summary,
            'daily_sales': daily_sales,
            'category_sales': category_sales,
            'best_sellers': best_sellers,
            'charts': charts
        }
        
        if format == 'pdf':
            return self._generate_pdf_report(report_data)
        elif format == 'excel':
            return self._generate_excel_report(report_data)
        elif format == 'html':
            return self._generate_html_report(report_data)
        else:
            raise ValueError(f"지원하지 않는 형식: {format}")
    
    def generate_inventory_report(self, format: str = 'pdf') -> str:
        """재고 리포트 생성"""
        conn = psycopg2.connect(**self.db_config)
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # 1. 재고 현황
        cursor.execute("""
            SELECT 
                category,
                COUNT(*) as product_count,
                SUM(stock) as total_stock,
                SUM(stock * price) as inventory_value,
                AVG(stock) as avg_stock
            FROM products
            GROUP BY category
            ORDER BY inventory_value DESC
        """)
        inventory_summary = cursor.fetchall()
        
        # 2. 재고 부족 상품
        cursor.execute("""
            SELECT 
                p.id,
                p.name,
                p.category,
                p.stock,
                COALESCE(AVG(oi.quantity), 0) as avg_daily_sales,
                CASE 
                    WHEN AVG(oi.quantity) > 0 THEN p.stock / AVG(oi.quantity)
                    ELSE 999
                END as days_remaining
            FROM products p
            LEFT JOIN order_items oi ON p.id = oi.product_id
            LEFT JOIN orders o ON oi.order_id = o.id 
                AND o.created_at >= CURRENT_DATE - INTERVAL '30 days'
            WHERE p.stock < 50
            GROUP BY p.id, p.name, p.category, p.stock
            HAVING p.stock / NULLIF(AVG(oi.quantity), 0) < 7
            ORDER BY days_remaining
            LIMIT 20
        """)
        low_stock_items = cursor.fetchall()
        
        # 3. 재고 회전율
        cursor.execute("""
            WITH inventory_turnover AS (
                SELECT 
                    p.category,
                    SUM(oi.quantity * oi.price) / NULLIF(AVG(p.stock * p.price), 0) as turnover_rate
                FROM products p
                LEFT JOIN order_items oi ON p.id = oi.product_id
                LEFT JOIN orders o ON oi.order_id = o.id
                WHERE o.created_at >= CURRENT_DATE - INTERVAL '90 days'
                GROUP BY p.category
            )
            SELECT * FROM inventory_turnover
            ORDER BY turnover_rate DESC
        """)
        turnover_rates = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        report_data = {
            'title': '재고 현황 리포트',
            'period': datetime.now().strftime('%Y-%m-%d'),
            'inventory_summary': inventory_summary,
            'low_stock_items': low_stock_items,
            'turnover_rates': turnover_rates
        }
        
        if format == 'pdf':
            return self._generate_pdf_report(report_data)
        elif format == 'excel':
            return self._generate_excel_report(report_data)
        else:
            raise ValueError(f"지원하지 않는 형식: {format}")
    
    def generate_customer_report(self, start_date: str, end_date: str, format: str = 'pdf') -> str:
        """고객 분석 리포트 생성"""
        conn = psycopg2.connect(**self.db_config)
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # 1. 고객 통계
        cursor.execute("""
            SELECT 
                COUNT(DISTINCT customer_id) as total_customers,
                COUNT(DISTINCT CASE WHEN order_count = 1 THEN customer_id END) as new_customers,
                COUNT(DISTINCT CASE WHEN order_count > 1 THEN customer_id END) as repeat_customers,
                AVG(total_spent) as avg_customer_value
            FROM (
                SELECT 
                    customer_id,
                    COUNT(*) as order_count,
                    SUM(total_amount) as total_spent
                FROM orders
                WHERE created_at BETWEEN %s AND %s
                GROUP BY customer_id
            ) customer_stats
        """, (start_date, end_date))
        customer_summary = cursor.fetchone()
        
        # 2. VIP 고객 (상위 10%)
        cursor.execute("""
            SELECT 
                customer_id,
                customer_name,
                COUNT(*) as order_count,
                SUM(total_amount) as total_spent,
                AVG(total_amount) as avg_order_value,
                MAX(created_at) as last_order_date
            FROM orders
            WHERE created_at BETWEEN %s AND %s
            GROUP BY customer_id, customer_name
            ORDER BY total_spent DESC
            LIMIT 20
        """, (start_date, end_date))
        vip_customers = cursor.fetchall()
        
        # 3. 고객 세그먼트 분석
        cursor.execute("""
            WITH customer_rfm AS (
                SELECT 
                    customer_id,
                    MAX(created_at) as last_order_date,
                    COUNT(*) as frequency,
                    SUM(total_amount) as monetary
                FROM orders
                WHERE created_at BETWEEN %s AND %s
                GROUP BY customer_id
            )
            SELECT 
                CASE 
                    WHEN last_order_date >= CURRENT_DATE - INTERVAL '30 days' THEN '활성'
                    WHEN last_order_date >= CURRENT_DATE - INTERVAL '90 days' THEN '휴면'
                    ELSE '이탈'
                END as segment,
                COUNT(*) as customer_count,
                AVG(frequency) as avg_frequency,
                AVG(monetary) as avg_monetary
            FROM customer_rfm
            GROUP BY segment
        """, (start_date, end_date))
        customer_segments = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        report_data = {
            'title': '고객 분석 리포트',
            'period': f"{start_date} ~ {end_date}",
            'customer_summary': customer_summary,
            'vip_customers': vip_customers,
            'customer_segments': customer_segments
        }
        
        if format == 'pdf':
            return self._generate_pdf_report(report_data)
        elif format == 'excel':
            return self._generate_excel_report(report_data)
        else:
            raise ValueError(f"지원하지 않는 형식: {format}")
    
    def _create_sales_charts(self, daily_sales: List[Dict], category_sales: List[Dict]) -> List[str]:
        """매출 차트 생성"""
        charts = []
        
        # 1. 일별 매출 추이 차트
        plt.figure(figsize=(10, 6))
        dates = [row['date'] for row in daily_sales]
        revenues = [row['revenue'] for row in daily_sales]
        
        plt.plot(dates, revenues, marker='o', linewidth=2, markersize=6)
        plt.title('일별 매출 추이', fontsize=16, fontweight='bold')
        plt.xlabel('날짜')
        plt.ylabel('매출액 (원)')
        plt.xticks(rotation=45)
        plt.grid(True, alpha=0.3)
        plt.tight_layout()
        
        chart_path = os.path.join(self.output_dir, f'daily_sales_{datetime.now().strftime("%Y%m%d%H%M%S")}.png')
        plt.savefig(chart_path, dpi=300, bbox_inches='tight')
        plt.close()
        charts.append(chart_path)
        
        # 2. 카테고리별 매출 파이 차트
        plt.figure(figsize=(8, 8))
        categories = [row['category'] for row in category_sales[:5]]  # 상위 5개
        revenues = [row['revenue'] for row in category_sales[:5]]
        
        plt.pie(revenues, labels=categories, autopct='%1.1f%%', startangle=90)
        plt.title('카테고리별 매출 비중', fontsize=16, fontweight='bold')
        plt.axis('equal')
        
        chart_path = os.path.join(self.output_dir, f'category_sales_{datetime.now().strftime("%Y%m%d%H%M%S")}.png')
        plt.savefig(chart_path, dpi=300, bbox_inches='tight')
        plt.close()
        charts.append(chart_path)
        
        return charts
    
    def _generate_pdf_report(self, report_data: Dict[str, Any]) -> str:
        """PDF 리포트 생성"""
        filename = f"{report_data['title'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
        filepath = os.path.join(self.output_dir, filename)
        
        doc = SimpleDocTemplate(filepath, pagesize=A4)
        story = []
        
        # 제목
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Title'],
            fontName='NanumGothic',
            fontSize=24,
            textColor=colors.HexColor('#1a1a1a'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        story.append(Paragraph(report_data['title'], title_style))
        story.append(Spacer(1, 20))
        
        # 기간
        period_style = ParagraphStyle(
            'Period',
            parent=self.styles['Normal'],
            fontName='NanumGothic',
            fontSize=12,
            alignment=TA_CENTER
        )
        story.append(Paragraph(f"기간: {report_data['period']}", period_style))
        story.append(Spacer(1, 30))
        
        # 요약 정보
        if 'summary' in report_data and report_data['summary']:
            summary = report_data['summary']
            summary_data = [
                ['항목', '값'],
                ['총 주문 수', f"{summary.get('total_orders', 0):,}건"],
                ['총 매출', f"₩{summary.get('total_revenue', 0):,.0f}"],
                ['평균 주문액', f"₩{summary.get('avg_order_value', 0):,.0f}"],
                ['고유 고객 수', f"{summary.get('unique_customers', 0):,}명"]
            ]
            
            summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, -1), 'NanumGothic'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(summary_table)
            story.append(Spacer(1, 30))
        
        # 차트 추가
        if 'charts' in report_data:
            for chart_path in report_data['charts']:
                if os.path.exists(chart_path):
                    img = Image(chart_path, width=6*inch, height=4*inch)
                    story.append(img)
                    story.append(Spacer(1, 20))
        
        # 상세 데이터 테이블들
        if 'best_sellers' in report_data and report_data['best_sellers']:
            story.append(Paragraph('베스트셀러 TOP 10', self.styles['Heading2']))
            story.append(Spacer(1, 10))
            
            best_sellers_data = [['순위', '상품명', '카테고리', '판매량', '매출']]
            for i, item in enumerate(report_data['best_sellers'], 1):
                best_sellers_data.append([
                    str(i),
                    item['name'][:30],
                    item['category'],
                    f"{item['quantity_sold']:,}개",
                    f"₩{item['revenue']:,.0f}"
                ])
            
            best_sellers_table = Table(best_sellers_data)
            best_sellers_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, -1), 'NanumGothic'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(best_sellers_table)
        
        # PDF 생성
        doc.build(story)
        
        logger.info(f"PDF 리포트 생성 완료: {filepath}")
        return filepath
    
    def _generate_excel_report(self, report_data: Dict[str, Any]) -> str:
        """Excel 리포트 생성"""
        filename = f"{report_data['title'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        
        wb = Workbook()
        
        # 요약 시트
        ws_summary = wb.active
        ws_summary.title = "요약"
        
        # 제목
        ws_summary['A1'] = report_data['title']
        ws_summary['A1'].font = Font(size=20, bold=True)
        ws_summary['A2'] = f"기간: {report_data['period']}"
        ws_summary['A2'].font = Font(size=12)
        
        # 요약 정보
        if 'summary' in report_data and report_data['summary']:
            summary = report_data['summary']
            ws_summary['A4'] = '항목'
            ws_summary['B4'] = '값'
            ws_summary['A5'] = '총 주문 수'
            ws_summary['B5'] = summary.get('total_orders', 0)
            ws_summary['A6'] = '총 매출'
            ws_summary['B6'] = summary.get('total_revenue', 0)
            ws_summary['A7'] = '평균 주문액'
            ws_summary['B7'] = summary.get('avg_order_value', 0)
            ws_summary['A8'] = '고유 고객 수'
            ws_summary['B8'] = summary.get('unique_customers', 0)
            
            # 스타일 적용
            for row in range(4, 9):
                ws_summary[f'A{row}'].font = Font(bold=True)
                ws_summary[f'B{row}'].number_format = '#,##0'
        
        # 일별 매출 시트
        if 'daily_sales' in report_data and report_data['daily_sales']:
            ws_daily = wb.create_sheet("일별 매출")
            headers = ['날짜', '주문 수', '매출']
            for col, header in enumerate(headers, 1):
                ws_daily.cell(row=1, column=col, value=header).font = Font(bold=True)
            
            for row, data in enumerate(report_data['daily_sales'], 2):
                ws_daily.cell(row=row, column=1, value=data['date'])
                ws_daily.cell(row=row, column=2, value=data['order_count'])
                ws_daily.cell(row=row, column=3, value=data['revenue']).number_format = '#,##0'
            
            # 차트 추가
            chart = LineChart()
            chart.title = "일별 매출 추이"
            chart.y_axis.title = "매출액"
            chart.x_axis.title = "날짜"
            
            data = Reference(ws_daily, min_col=3, min_row=1, max_row=len(report_data['daily_sales'])+1)
            categories = Reference(ws_daily, min_col=1, min_row=2, max_row=len(report_data['daily_sales'])+1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            ws_daily.add_chart(chart, "E2")
        
        # 베스트셀러 시트
        if 'best_sellers' in report_data and report_data['best_sellers']:
            ws_best = wb.create_sheet("베스트셀러")
            headers = ['순위', '상품 ID', '상품명', '카테고리', '판매량', '매출']
            for col, header in enumerate(headers, 1):
                ws_best.cell(row=1, column=col, value=header).font = Font(bold=True)
            
            for row, item in enumerate(report_data['best_sellers'], 2):
                ws_best.cell(row=row, column=1, value=row-1)
                ws_best.cell(row=row, column=2, value=item['id'])
                ws_best.cell(row=row, column=3, value=item['name'])
                ws_best.cell(row=row, column=4, value=item['category'])
                ws_best.cell(row=row, column=5, value=item['quantity_sold'])
                ws_best.cell(row=row, column=6, value=item['revenue']).number_format = '#,##0'
        
        # 파일 저장
        wb.save(filepath)
        
        logger.info(f"Excel 리포트 생성 완료: {filepath}")
        return filepath
    
    def _generate_html_report(self, report_data: Dict[str, Any]) -> str:
        """HTML 리포트 생성"""
        template_dir = os.path.join(os.path.dirname(__file__), 'templates')
        os.makedirs(template_dir, exist_ok=True)
        
        # HTML 템플릿 생성
        html_template = """
<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{{ title }}</title>
    <style>
        body {
            font-family: 'Nanum Gothic', Arial, sans-serif;
            margin: 0;
            padding: 20px;
            background-color: #f5f5f5;
        }
        .container {
            max-width: 1200px;
            margin: 0 auto;
            background-color: white;
            padding: 30px;
            border-radius: 10px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }
        h1 {
            color: #333;
            text-align: center;
            margin-bottom: 30px;
        }
        .period {
            text-align: center;
            color: #666;
            margin-bottom: 40px;
        }
        .summary {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
            gap: 20px;
            margin-bottom: 40px;
        }
        .summary-item {
            background-color: #f8f9fa;
            padding: 20px;
            border-radius: 8px;
            text-align: center;
        }
        .summary-item h3 {
            margin: 0 0 10px 0;
            color: #666;
            font-size: 14px;
        }
        .summary-item .value {
            font-size: 24px;
            font-weight: bold;
            color: #333;
        }
        table {
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
        }
        th, td {
            padding: 12px;
            text-align: left;
            border-bottom: 1px solid #ddd;
        }
        th {
            background-color: #f8f9fa;
            font-weight: bold;
        }
        tr:hover {
            background-color: #f5f5f5;
        }
        .chart {
            margin: 20px 0;
            text-align: center;
        }
        .chart img {
            max-width: 100%;
            height: auto;
            border-radius: 8px;
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>{{ title }}</h1>
        <p class="period">기간: {{ period }}</p>
        
        {% if summary %}
        <div class="summary">
            <div class="summary-item">
                <h3>총 주문 수</h3>
                <div class="value">{{ "{:,}".format(summary.total_orders) }}건</div>
            </div>
            <div class="summary-item">
                <h3>총 매출</h3>
                <div class="value">₩{{ "{:,.0f}".format(summary.total_revenue) }}</div>
            </div>
            <div class="summary-item">
                <h3>평균 주문액</h3>
                <div class="value">₩{{ "{:,.0f}".format(summary.avg_order_value) }}</div>
            </div>
            <div class="summary-item">
                <h3>고유 고객 수</h3>
                <div class="value">{{ "{:,}".format(summary.unique_customers) }}명</div>
            </div>
        </div>
        {% endif %}
        
        {% if charts %}
        <div class="charts">
            {% for chart in charts %}
            <div class="chart">
                <img src="{{ chart }}" alt="차트">
            </div>
            {% endfor %}
        </div>
        {% endif %}
        
        {% if best_sellers %}
        <h2>베스트셀러 TOP 10</h2>
        <table>
            <thead>
                <tr>
                    <th>순위</th>
                    <th>상품명</th>
                    <th>카테고리</th>
                    <th>판매량</th>
                    <th>매출</th>
                </tr>
            </thead>
            <tbody>
                {% for item in best_sellers %}
                <tr>
                    <td>{{ loop.index }}</td>
                    <td>{{ item.name }}</td>
                    <td>{{ item.category }}</td>
                    <td>{{ "{:,}".format(item.quantity_sold) }}개</td>
                    <td>₩{{ "{:,.0f}".format(item.revenue) }}</td>
                </tr>
                {% endfor %}
            </tbody>
        </table>
        {% endif %}
    </div>
</body>
</html>
        """
        
        # 템플릿 파일 저장
        with open(os.path.join(template_dir, 'report.html'), 'w', encoding='utf-8') as f:
            f.write(html_template)
        
        # Jinja2 환경 설정
        env = Environment(loader=FileSystemLoader(template_dir))
        template = env.get_template('report.html')
        
        # HTML 렌더링
        html_content = template.render(**report_data)
        
        # 파일 저장
        filename = f"{report_data['title'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d%H%M%S')}.html"
        filepath = os.path.join(self.output_dir, filename)
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        logger.info(f"HTML 리포트 생성 완료: {filepath}")
        return filepath